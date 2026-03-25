"""파일 관련 ACTION 핸들러: OPEN_FILE, SAVE_FILE."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from autooffice.engine.actions.base import ActionHandler
from autooffice.engine.context import EngineContext
from autooffice.models.action_result import ActionResult

logger = logging.getLogger(__name__)


class OpenFileHandler(ActionHandler):
    """OPEN_FILE: 엑셀 파일을 열어 컨텍스트에 등록한다.

    params:
        file_path: 파일 경로 (상대경로면 data_dir 기준)
        alias: 컨텍스트 등록명 (선택, 기본값은 파일명)
        data_only: True면 수식 대신 값만 읽기 (기본 False)
    """

    def execute(self, params: dict[str, Any], ctx: EngineContext) -> ActionResult:
        file_path = params.get("file_path", "")
        alias = params.get("alias", "")
        data_only = params.get("data_only", False)

        path = Path(file_path)
        if not path.is_absolute():
            path = ctx.data_dir / path

        if not path.exists():
            return ActionResult(
                success=False,
                error=f"파일을 찾을 수 없습니다: {path}",
            )

        try:
            wb = load_workbook(str(path), data_only=data_only)
            name = alias or path.stem
            ctx.register_workbook(name, wb, path)
            return ActionResult(
                success=True,
                data={"sheets": wb.sheetnames},
                message=f"파일 열기 완료: {path.name} (시트: {wb.sheetnames})",
            )
        except Exception as e:
            return ActionResult(success=False, error=f"파일 열기 실패: {e}")


class SaveFileHandler(ActionHandler):
    """SAVE_FILE: 워크북을 저장한다.

    params:
        file: 컨텍스트에 등록된 워크북 alias
        save_as: 다른 이름으로 저장 (선택)
    """

    def execute(self, params: dict[str, Any], ctx: EngineContext) -> ActionResult:
        alias = params.get("file", "")
        save_as = params.get("save_as", "")

        try:
            wb = ctx.get_workbook(alias)
            if save_as:
                save_path = Path(save_as)
                if not save_path.is_absolute():
                    save_path = ctx.data_dir / save_path
            else:
                save_path = ctx.file_paths[alias]

            wb.save(str(save_path))
            return ActionResult(
                success=True,
                message=f"파일 저장 완료: {save_path.name}",
            )
        except Exception as e:
            return ActionResult(success=False, error=f"파일 저장 실패: {e}")
