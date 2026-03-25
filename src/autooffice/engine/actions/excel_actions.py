"""엑셀 조작 ACTION 핸들러: READ_COLUMNS, READ_RANGE, WRITE_DATA, CLEAR_RANGE, RECALCULATE."""

from __future__ import annotations

import logging
from typing import Any

from openpyxl.utils import get_column_letter, column_index_from_string

from autooffice.engine.actions.base import ActionHandler
from autooffice.engine.context import EngineContext
from autooffice.models.action_result import ActionResult

logger = logging.getLogger(__name__)


class ReadColumnsHandler(ActionHandler):
    """READ_COLUMNS: 특정 시트에서 지정 컬럼의 데이터를 읽는다.

    params:
        file: 워크북 alias
        sheet: 시트명
        columns: 읽을 컬럼 목록 (헤더명 또는 컬럼 문자)
        start_row: 시작 행 (기본 2, 헤더 다음)
        end_row: 종료 행 (선택, 미지정 시 데이터 끝까지)
    """

    def execute(self, params: dict[str, Any], ctx: EngineContext) -> ActionResult:
        alias = params.get("file", "")
        sheet_name = params.get("sheet", "")
        columns = params.get("columns", [])
        start_row = params.get("start_row", 2)
        end_row = params.get("end_row")

        try:
            wb = ctx.get_workbook(alias)
            ws = wb[sheet_name]

            # 헤더 행에서 컬럼 인덱스 매핑
            header_map = {}
            for cell in ws[1]:
                if cell.value is not None:
                    header_map[str(cell.value)] = cell.column

            # 컬럼 인덱스 결정
            col_indices = []
            for col in columns:
                if col in header_map:
                    col_indices.append(header_map[col])
                elif isinstance(col, str) and col.isalpha() and col.isascii():
                    col_indices.append(column_index_from_string(col))
                else:
                    return ActionResult(
                        success=False,
                        error=f"컬럼 '{col}'을 찾을 수 없습니다. 사용 가능: {list(header_map.keys())}",
                    )

            # 데이터 읽기
            actual_end = end_row or ws.max_row
            data = []
            for row_idx in range(start_row, actual_end + 1):
                row_data = {}
                for col_name, col_idx in zip(columns, col_indices):
                    row_data[col_name] = ws.cell(row=row_idx, column=col_idx).value
                # 빈 행 스킵 (모든 값이 None)
                if any(v is not None for v in row_data.values()):
                    data.append(row_data)

            return ActionResult(
                success=True,
                data=data,
                message=f"{len(data)}행 읽기 완료 (컬럼: {columns})",
            )
        except Exception as e:
            return ActionResult(success=False, error=f"READ_COLUMNS 실패: {e}")


class ReadRangeHandler(ActionHandler):
    """READ_RANGE: 셀 범위의 데이터를 읽는다.

    params:
        file: 워크북 alias
        sheet: 시트명
        range: 셀 범위 (예: "B3:F100")
    """

    def execute(self, params: dict[str, Any], ctx: EngineContext) -> ActionResult:
        alias = params.get("file", "")
        sheet_name = params.get("sheet", "")
        cell_range = params.get("range", "")

        try:
            wb = ctx.get_workbook(alias)
            ws = wb[sheet_name]

            data = []
            for row in ws[cell_range]:
                row_values = [cell.value for cell in row]
                data.append(row_values)

            return ActionResult(
                success=True,
                data=data,
                message=f"범위 {cell_range} 읽기 완료 ({len(data)}행)",
            )
        except Exception as e:
            return ActionResult(success=False, error=f"READ_RANGE 실패: {e}")


class WriteDataHandler(ActionHandler):
    """WRITE_DATA: 데이터를 양식의 지정 위치에 쓴다.

    params:
        source: 데이터 소스 ($변수명 참조 또는 직접 리스트)
        target_file: 대상 워크북 alias
        target_sheet: 대상 시트명
        target_start: 시작 셀 (예: "B3")
        column_mapping: 소스 컬럼 → 대상 열 매핑 (선택)
            예: {"원본컬럼A": "B", "원본컬럼B": "C"}
    """

    def execute(self, params: dict[str, Any], ctx: EngineContext) -> ActionResult:
        source = params.get("source", [])
        target_file = params.get("target_file", "")
        target_sheet = params.get("target_sheet", "")
        target_start = params.get("target_start", "A1")
        column_mapping = params.get("column_mapping")

        try:
            wb = ctx.get_workbook(target_file)
            ws = wb[target_sheet]

            # 시작 셀 파싱
            from openpyxl.utils.cell import coordinate_from_string
            col_letter, start_row = coordinate_from_string(target_start)
            start_col = column_index_from_string(col_letter)

            if not isinstance(source, list) or len(source) == 0:
                return ActionResult(
                    success=False,
                    error="source 데이터가 비어있거나 리스트가 아닙니다.",
                )

            rows_written = 0
            for i, row_data in enumerate(source):
                current_row = start_row + i

                if isinstance(row_data, dict):
                    if column_mapping:
                        # 매핑 기반 쓰기
                        for src_col, tgt_col_letter in column_mapping.items():
                            tgt_col_idx = column_index_from_string(tgt_col_letter)
                            ws.cell(
                                row=current_row,
                                column=tgt_col_idx,
                                value=row_data.get(src_col),
                            )
                    else:
                        # 순서대로 쓰기
                        for j, value in enumerate(row_data.values()):
                            ws.cell(
                                row=current_row,
                                column=start_col + j,
                                value=value,
                            )
                elif isinstance(row_data, (list, tuple)):
                    for j, value in enumerate(row_data):
                        ws.cell(
                            row=current_row,
                            column=start_col + j,
                            value=value,
                        )

                rows_written += 1

            return ActionResult(
                success=True,
                data={"rows_written": rows_written},
                message=f"{rows_written}행 쓰기 완료 → {target_sheet}!{target_start}",
            )
        except Exception as e:
            return ActionResult(success=False, error=f"WRITE_DATA 실패: {e}")


class ClearRangeHandler(ActionHandler):
    """CLEAR_RANGE: 지정 범위의 셀 값을 비운다.

    params:
        file: 워크북 alias
        sheet: 시트명
        range: 셀 범위 (예: "B3:F100")
    """

    def execute(self, params: dict[str, Any], ctx: EngineContext) -> ActionResult:
        alias = params.get("file", "")
        sheet_name = params.get("sheet", "")
        cell_range = params.get("range", "")

        try:
            wb = ctx.get_workbook(alias)
            ws = wb[sheet_name]

            cells_cleared = 0
            for row in ws[cell_range]:
                for cell in row:
                    if cell.value is not None:
                        cell.value = None
                        cells_cleared += 1

            return ActionResult(
                success=True,
                data={"cells_cleared": cells_cleared},
                message=f"범위 {cell_range} 클리어 완료 ({cells_cleared}셀)",
            )
        except Exception as e:
            return ActionResult(success=False, error=f"CLEAR_RANGE 실패: {e}")


class RecalculateHandler(ActionHandler):
    """RECALCULATE: 워크북의 수식 재계산을 트리거한다.

    params:
        file: 워크북 alias

    Note: openpyxl은 수식 재계산을 직접 수행하지 못한다.
    이 핸들러는 워크북을 저장하여 Excel에서 열 때 재계산되도록 한다.
    수식 결과를 프로그래밍적으로 계산하려면 xlcalc 등 별도 라이브러리가 필요하다.
    """

    def execute(self, params: dict[str, Any], ctx: EngineContext) -> ActionResult:
        alias = params.get("file", "")

        try:
            wb = ctx.get_workbook(alias)
            # calcChain 초기화 → Excel이 열 때 전체 재계산
            wb.calculation.calcMode = "auto"
            return ActionResult(
                success=True,
                message=f"재계산 플래그 설정 완료: {alias} (Excel 열 때 재계산됨)",
            )
        except Exception as e:
            return ActionResult(success=False, error=f"RECALCULATE 실패: {e}")
