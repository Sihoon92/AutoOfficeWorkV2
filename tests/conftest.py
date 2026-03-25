"""테스트 공통 fixture."""

from __future__ import annotations

import json
from pathlib import Path

import pytest
from openpyxl import Workbook

from autooffice.engine.actions import build_default_registry
from autooffice.engine.context import EngineContext
from autooffice.engine.runner import PlanRunner
from autooffice.models.execution_plan import ExecutionPlan


@pytest.fixture
def tmp_data_dir(tmp_path: Path) -> Path:
    """임시 데이터 디렉토리."""
    return tmp_path


@pytest.fixture
def engine_ctx(tmp_data_dir: Path) -> EngineContext:
    """기본 EngineContext."""
    return EngineContext(data_dir=tmp_data_dir)


@pytest.fixture
def runner() -> PlanRunner:
    """기본 PlanRunner (모든 핸들러 등록)."""
    return PlanRunner(build_default_registry())


@pytest.fixture
def sample_raw_excel(tmp_data_dir: Path) -> Path:
    """더미 raw 데이터 엑셀 파일 생성."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 헤더
    headers = ["날짜", "라인명", "생산수량", "불량수량", "불량유형"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # 더미 데이터 (5행)
    data = [
        ["2026-03-25", "코팅1라인", 1000, 30, "외관"],
        ["2026-03-25", "코팅2라인", 800, 20, "치수"],
        ["2026-03-25", "조립1라인", 1200, 80, "기능"],
        ["2026-03-25", "조립2라인", 900, 15, "외관"],
        ["2026-03-25", "포장라인", 1500, 10, "기타"],
    ]
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    path = tmp_data_dir / "raw_data.xlsx"
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def sample_template_excel(tmp_data_dir: Path) -> Path:
    """더미 양식 엑셀 파일 생성."""
    wb = Workbook()

    # 데이터입력 시트
    ws_input = wb.active
    ws_input.title = "데이터입력"
    for col, header in enumerate(["날짜", "라인명", "생산수량", "불량수량", "불량유형"], 2):
        ws_input.cell(row=1, column=col, value=header)

    # 일별현황 시트
    ws_daily = wb.create_sheet("일별현황")
    daily_headers = ["날짜", "라인명", "생산수량", "불량수량", "불량유형", "불량률"]
    for col, header in enumerate(daily_headers, 1):
        ws_daily.cell(row=1, column=col, value=header)

    # 주별현황, 월별현황 시트
    wb.create_sheet("주별현황")
    wb.create_sheet("월별현황")

    path = tmp_data_dir / "defect_rate_template.xlsx"
    wb.save(str(path))
    wb.close()
    return path


@pytest.fixture
def sample_plan_dict() -> dict:
    """더미 execution_plan JSON dict."""
    return {
        "task_id": "test_defect_rate",
        "description": "테스트용 불량률 산출",
        "created_by": "Claude",
        "created_at": "2026-03-25T09:00:00+09:00",
        "version": "1.0.0",
        "metadata": {
            "task_type": "defect_rate",
            "template_hash": "test_hash_123",
            "reusable": True,
        },
        "inputs": {
            "raw_data": {
                "description": "생산 데이터",
                "expected_format": "xlsx",
                "expected_sheets": ["Sheet1"],
                "expected_columns": ["날짜", "라인명", "생산수량", "불량수량", "불량유형"],
            },
            "template": {
                "description": "불량률 양식",
                "expected_format": "xlsx",
                "expected_sheets": ["데이터입력", "일별현황"],
            },
        },
        "steps": [
            {
                "step": 1,
                "action": "OPEN_FILE",
                "description": "raw 데이터 열기",
                "params": {"file_path": "raw_data.xlsx", "alias": "raw", "data_only": True},
                "on_fail": "STOP",
            },
            {
                "step": 2,
                "action": "OPEN_FILE",
                "description": "양식 열기",
                "params": {"file_path": "defect_rate_template.xlsx", "alias": "template"},
                "on_fail": "STOP",
            },
            {
                "step": 3,
                "action": "READ_COLUMNS",
                "description": "raw 데이터 읽기",
                "params": {
                    "file": "raw",
                    "sheet": "Sheet1",
                    "columns": ["날짜", "라인명", "생산수량", "불량수량", "불량유형"],
                    "start_row": 2,
                },
                "store_as": "raw_data",
                "on_fail": "STOP",
            },
            {
                "step": 4,
                "action": "CLEAR_RANGE",
                "description": "기존 데이터 클리어",
                "params": {"file": "template", "sheet": "데이터입력", "range": "B3:F100"},
                "on_fail": "STOP",
            },
            {
                "step": 5,
                "action": "WRITE_DATA",
                "description": "데이터 쓰기",
                "params": {
                    "source": "$raw_data",
                    "target_file": "template",
                    "target_sheet": "데이터입력",
                    "target_start": "B3",
                    "column_mapping": {
                        "날짜": "B",
                        "라인명": "C",
                        "생산수량": "D",
                        "불량수량": "E",
                        "불량유형": "F",
                    },
                },
                "on_fail": "STOP",
            },
            {
                "step": 6,
                "action": "SAVE_FILE",
                "description": "결과 저장",
                "params": {"file": "template", "save_as": "result.xlsx"},
                "on_fail": "STOP",
            },
            {
                "step": 7,
                "action": "LOG",
                "description": "완료 로그",
                "params": {"message": "테스트 완료", "level": "info"},
                "on_fail": "SKIP",
            },
        ],
        "final_output": {"type": "file", "description": "결과 엑셀 파일"},
    }
