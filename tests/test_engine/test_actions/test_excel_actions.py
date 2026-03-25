"""엑셀 ACTION 핸들러 테스트."""

from __future__ import annotations

from openpyxl import Workbook

from autooffice.engine.actions.excel_actions import (
    ClearRangeHandler,
    ReadColumnsHandler,
    WriteDataHandler,
)
from autooffice.engine.context import EngineContext


class TestReadColumns:
    def test_read_by_header_name(self, engine_ctx: EngineContext, sample_raw_excel):
        """헤더 이름으로 컬럼 읽기."""
        from openpyxl import load_workbook

        wb = load_workbook(str(sample_raw_excel))
        engine_ctx.register_workbook("raw", wb, sample_raw_excel)

        handler = ReadColumnsHandler()
        result = handler.execute(
            {
                "file": "raw",
                "sheet": "Sheet1",
                "columns": ["날짜", "라인명", "생산수량"],
                "start_row": 2,
            },
            engine_ctx,
        )

        assert result.success
        assert len(result.data) == 5
        assert "날짜" in result.data[0]
        assert "라인명" in result.data[0]
        assert result.data[0]["라인명"] == "코팅1라인"

    def test_read_missing_column(self, engine_ctx: EngineContext, sample_raw_excel):
        """존재하지 않는 컬럼 읽기 시 실패."""
        from openpyxl import load_workbook

        wb = load_workbook(str(sample_raw_excel))
        engine_ctx.register_workbook("raw", wb, sample_raw_excel)

        handler = ReadColumnsHandler()
        result = handler.execute(
            {
                "file": "raw",
                "sheet": "Sheet1",
                "columns": ["존재하지않는컬럼"],
                "start_row": 2,
            },
            engine_ctx,
        )

        assert not result.success
        assert "찾을 수 없습니다" in result.error


class TestWriteData:
    def test_write_with_column_mapping(self, engine_ctx: EngineContext, tmp_data_dir):
        """컬럼 매핑으로 데이터 쓰기."""
        wb = Workbook()
        ws = wb.active
        ws.title = "데이터"
        engine_ctx.register_workbook("target", wb, tmp_data_dir / "target.xlsx")

        source_data = [
            {"이름": "A", "값": 100},
            {"이름": "B", "값": 200},
        ]

        handler = WriteDataHandler()
        result = handler.execute(
            {
                "source": source_data,
                "target_file": "target",
                "target_sheet": "데이터",
                "target_start": "B3",
                "column_mapping": {"이름": "B", "값": "C"},
            },
            engine_ctx,
        )

        assert result.success
        assert result.data["rows_written"] == 2
        assert ws.cell(row=3, column=2).value == "A"
        assert ws.cell(row=3, column=3).value == 100
        assert ws.cell(row=4, column=2).value == "B"
        assert ws.cell(row=4, column=3).value == 200


class TestClearRange:
    def test_clear_range(self, engine_ctx: EngineContext, tmp_data_dir):
        """셀 범위 클리어."""
        wb = Workbook()
        ws = wb.active
        ws.title = "시트"
        ws.cell(row=1, column=1, value="헤더")
        ws.cell(row=2, column=1, value="데이터1")
        ws.cell(row=3, column=1, value="데이터2")
        engine_ctx.register_workbook("test", wb, tmp_data_dir / "test.xlsx")

        handler = ClearRangeHandler()
        result = handler.execute(
            {"file": "test", "sheet": "시트", "range": "A2:A3"},
            engine_ctx,
        )

        assert result.success
        assert ws.cell(row=1, column=1).value == "헤더"  # 보존
        assert ws.cell(row=2, column=1).value is None  # 클리어됨
        assert ws.cell(row=3, column=1).value is None  # 클리어됨
